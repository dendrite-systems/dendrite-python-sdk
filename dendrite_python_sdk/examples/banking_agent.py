from datetime import datetime, timedelta
import logging
import os
import asyncio
from typing import List
from pydantic import BaseModel
import openpyxl
from dotenv import load_dotenv, find_dotenv
from openai import OpenAI

from dendrite_python_sdk.dendrite_browser.DendriteBrowser import DendriteBrowser

load_dotenv(find_dotenv())
logging.getLogger().setLevel(logging.ERROR)


class AccountDetails(BaseModel):
    iban: str
    bic: str
    adress: str
    balance: float


async def list_account_information() -> AccountDetails:
    dendrite = DendriteBrowser()

    # Download the Dendrite Vault plugin and authenticate on the website you want your agent to be able to access
    await dendrite.authenticate("seb.se")

    # Here we navigate to the dashboard and specify `expected_page` so that an exception will be thrown if we aren't logged in
    page = await dendrite.goto(
        "https://apps.seb.se/ccs/ibf/kgb/1000/1100/kgbc1101.aspx",
        expected_page="Should be the bank's dashboard and not a log in page. This can happen if we aren't authenticated.",
    )

    # Next we navigate through a couple of tabs
    el = await page.get_element("the företagskonto link")
    await el.click()

    el = await page.get_element("the kontoinformation tab")
    await el.click(expected_outcome="See the details about the account")

    # Now we extract all the account details as specifed in the AccountDetails model and return them
    res = await page.extract("Please extract the account details", AccountDetails)

    return res.return_data


def get_file_path():
    # Define the project root and test directory
    filepath = os.path.join(
        "test", f"transactions_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )

    # Check if the file already exists
    if os.path.exists(filepath):
        # Get the file's last modification time
        file_mod_time = datetime.fromtimestamp(os.path.getmtime(filepath))

        # Check if the file is less than a day old
        if datetime.now() - file_mod_time < timedelta(days=1):
            # Read the Excel file and count the rows
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            row_count = sheet.max_row

            # Return the filepath if the cache is valid
            return filepath

    return filepath


async def get_transactions() -> str:
    filepath = os.path.join(
        "test", f"transactions_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )
    if os.path.exists(filepath):
        return filepath

    dendrite = DendriteBrowser()

    # Attempt authentication and navigate to dashboard
    await dendrite.authenticate("seb.se")
    page = await dendrite.goto(
        "https://apps.seb.se/ccs/ibf/kgb/1000/1100/kgbc1101.aspx",
        expected_page="Should be the bank's dashboard and not a log in page. This can happen if we aren't authenticated.",
    )

    # Dashboard takes a while to load, wait until it is done and press the main bank accounts link to visit it
    await page.wait_for("Wait until the 'företagskonto' link is available.")
    el = await page.get_element("the företagskonto link")
    await el.click(
        expected_outcome="Should redirect to the account's page with listed transactions"
    )

    # Open modal to download all transaction data
    el = await page.get_element("the Spara kontohändelser")
    await el.click(
        expected_outcome="A modal should show up with the option save the transactions as e.g csv"
    )

    # Press the download button to downloads
    el = await page.get_element("the 'Spara' button in the download modal")
    await el.click()

    # Get the downloaded transactions
    download = await page.get_download()

    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    # Save the downloaded file to the specified path
    await download.save_as(filepath)

    # Read the Excel file and count the rows
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active
    row_count = sheet.max_row

    res = f"Downloaded transactions to {filepath}, amount of rows are: {row_count}"
    # print(res)
    return filepath


load_dotenv(find_dotenv())

client = OpenAI()


async def analyze_transactions(prompt: str):
    # First, download the transactions
    filepath = await get_transactions()

    file = client.files.create(file=open(filepath, "rb"), purpose="assistants")

    # Create a new assistant with the file attached and code interpreter enabled
    analysis_assistant = client.beta.assistants.create(
        name="Transaction Analysis Assistant",
        instructions="""Analyze the transactions in the uploaded Excel file based on the user's prompt. 
        Use the code interpreter to perform calculations and data analysis as needed. 
        Provide insights, patterns, and financial advice as appropriate.""",
        model="gpt-4o",
        tools=[{"type": "code_interpreter"}],
        tool_resources={"code_interpreter": {"file_ids": [file.id]}},
    )
    thread = client.beta.threads.create(
        messages=[
            {
                "role": "user",
                "content": prompt,
                "attachments": [
                    {"file_id": file.id, "tools": [{"type": "code_interpreter"}]}
                ],
            }
        ]
    )
    run = client.beta.threads.runs.create_and_poll(
        thread_id=thread.id,
        assistant_id=analysis_assistant.id,
    )

    if run.status == "completed":
        messages = client.beta.threads.messages.list(thread_id=thread.id)
        # print("messages: ", messages)
        return str(messages)
    else:
        pass  # print(run.status)


# Define the functions that the assistant can call
functions = [
    {
        "name": "list_account_information",
        "description": "Get account details including Bankgiro, IBAN, BIC, Address, and Balance",
        "parameters": {"type": "object", "properties": {}},
    },
    {
        "name": "get_transactions",
        "description": "Download recent transactions and return the file path and row count",
        "parameters": {"type": "object", "properties": {}},
    },
    {
        "name": "analyze_transactions",
        "description": "Download and analyze transactions based on a given prompt",
        "parameters": {
            "type": "object",
            "properties": {
                "prompt": {
                    "type": "string",
                    "description": "Specific instructions for transaction analysis",
                }
            },
            "required": ["prompt"],
        },
    },
]

# Create an assistant
assistant = client.beta.assistants.create(
    name="Bank Assistant",
    instructions="""You are a helpful assistant for banking operations. You can retrieve messages, list account information, get transactions, and analyze them. Use the available functions when necessary to assist the user.

When analyzing transactions, always use the analyze_transactions function, which will automatically download the latest transactions before analysis.

If a bank function fails to authenticate, suggest the user to use the Dendrite Vault plugin to authenticate.

Whenever a function fails to run, list the entire error or expections so I can debug the function.

Provide clear and concise responses, and offer insights based on the banking data when appropriate.""",
    model="gpt-4o",
    tools=[{"type": "function", "function": func} for func in functions],  # type: ignore
)

# Create a thread for the main conversation
thread = client.beta.threads.create()


async def run_conversation():
    print("\n\n\nSEB Banking Agent. Send 'exit' to exit the chat.")
    while True:
        user_input = input("\n\nYou: ")
        if user_input.lower() == "exit":
            print("Assistant: Goodbye!")
            break

        # Add the user's message to the thread
        client.beta.threads.messages.create(
            thread_id=thread.id, role="user", content=user_input
        )

        # Run the assistant
        run = client.beta.threads.runs.create(
            thread_id=thread.id, assistant_id=assistant.id
        )

        # Wait for the run to complete
        while run.status != "completed":
            run = client.beta.threads.runs.retrieve(thread_id=thread.id, run_id=run.id)
            if run.status == "requires_action":
                for tool_call in run.required_action.submit_tool_outputs.tool_calls:  # type: ignore
                    function_name = tool_call.function.name
                    function_args = eval(tool_call.function.arguments)

                    # Call the appropriate function
                    try:
                        if function_name == "list_account_information":
                            result = await list_account_information()
                        elif function_name == "get_transactions":
                            result = await get_transactions()
                        elif function_name == "analyze_transactions":
                            result = await analyze_transactions(function_args["prompt"])
                    except Exception as e:
                        result = str(e)
                        print("\n\nException: ", result)

                    # Submit the function result back to the assistant
                    client.beta.threads.runs.submit_tool_outputs(
                        thread_id=thread.id,
                        run_id=run.id,
                        tool_outputs=[
                            {"tool_call_id": tool_call.id, "output": str(result)}  # type: ignore
                        ],
                    )
            await asyncio.sleep(1)

        # Retrieve and print the assistant's response
        messages = client.beta.threads.messages.list(thread_id=thread.id)
        assistant_messages = [msg for msg in messages if msg.role == "assistant"]
        if assistant_messages:
            print("\n\nAssistant:", assistant_messages[0].content[0].text.value)  # type: ignore


if __name__ == "__main__":
    # asyncio.run(list_account_information())
    asyncio.run(run_conversation())
